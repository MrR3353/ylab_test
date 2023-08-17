import asyncio
import functools
import os
from typing import AsyncGenerator, Callable

import openpyxl
from httpx import AsyncClient
from sqlalchemy import NullPool
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine
from sqlalchemy.orm import sessionmaker

from api.models import metadata
from api.schemas import (
    DishRequest,
    DishResponse,
    MenuDetailsResponse,
    MenuRequest,
    MenuResponse,
    SubmenuDetailsResponse,
    SubmenuRequest,
    SubmenuResponse,
)
from cache.cache_repository import CacheEntity as ce
from cache.cache_repository import CacheRepository
from config import (
    CELERY_ON,
    CELERY_TASK_TIME,
    DB_HOST,
    DB_NAME,
    DB_PASS,
    DB_PORT,
    DB_USER,
    get_project_root,
)
from database import get_async_session
from main import app
from tasks.celery_worker import celery

FILENAME = 'Menu.xlsx'
FILE_PATH = os.path.join(get_project_root(), 'admin', FILENAME)

# DATABASE  override for assigning new
DATABASE_URL = f'postgresql+asyncpg://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}'

engine_test = create_async_engine(DATABASE_URL, poolclass=NullPool, echo=False)
async_session_maker = sessionmaker(engine_test, class_=AsyncSession, expire_on_commit=False)
metadata.bind = engine_test

cache = CacheRepository()


async def override_get_async_session() -> AsyncGenerator[AsyncSession, None]:
    async with async_session_maker() as session:
        yield session

app.dependency_overrides[get_async_session] = override_get_async_session


async def ac() -> AsyncGenerator[AsyncClient, None]:
    async with AsyncClient(app=app, base_url='http://test') as client:
        yield client


def async_client(func: Callable):
    @functools.wraps(func)
    async def wrapper(*args, **kwargs):
        client = await ac().__anext__()
        return await func(*args, client, **kwargs)
    return wrapper


def get_excel_menu() -> list[MenuDetailsResponse]:
    '''
    Problem: submenus and dishes from xlsx file can have the same id.

    Solution 1: (Without editing xlsx file, but with a limited number of menu, submenu, dishes)
    Create id for submenu concatenating with menu_id like strings: menu_id + submenu_id
    Create id for dish like this: menu_id + submenu_id + dish_id
    But it won't work in this case:
    -------------------------------------------------
    Menu    Submenu         Dish        dish_id
    1
                1
                            1
                            2
                            ..
                            11      1 + 1 + 11 = 1111
                ..
                11
                            1       1 + 11 + 1 = 1111
    -------------------------------------------------
    To fix this we can use something like decimal conversion algorithm.
    (X - base of the number system)
    menu_id * X^2 + submenu_id * X + dish_id
    Example:
    1 * 100 + 1 * 10 + 11 * 1 = 121
    1 * 100 + 11 * 10 + 1 * 1 = 211
    But if X <= id, it won't work, example:
    1 * 100 + 0 * 10 + 0 * 1 = 100
    0 * 100 + 10 * 10 + 0 * 1 == 100
    So, I calculated max X (base of the number system), we can use in program.
    Max number value in this solution =
    (x - 1) * x^2 + (x - 1) * x + (x - 1) = x^3 - x^2 + x^2 - x + x - 1 = x^3 - 1
    x^3 - 1 = 2147483647 (max integer)
    So, max X = 1290, and that means, we can't use more than 1289 menu, submenu or dishes

    Solution 2: (editing xlsx file)
    Just count number of submenu, dishes while reading the file and change submenu_id, dishes_id
    Problem: if I try to add menu/submenu/dish with specific id, then will be problem because
    when endpoint will add new menu/submenu/dish without id with autoincrement if in db already contains
    menu/submenu/dish with such id as autoincrement choose, we get error: sqlalchemy.exc.IntegrityError:
     (sqlalchemy.dialects.postgresql.asyncpg.IntegrityError) <class 'asyncpg.exceptions.UniqueViolationError'>:
      повторяющееся значение ключа нарушает ограничение уникальности "menu_pkey"
      Can be fixed using uuid

    Solution 3: (editing xlsx file)   <- used this
    Just read excel, and if in db not such menu/submenu/dish with specified id, we add new record in db,
     get id for these record and change id in xlsx file. If db contains record with specified id in xlsx file,
    check that they equal, if not update record in db. If db contains records, which not specified in xlsx file
    then delete them
    '''

    try:
        workbook = openpyxl.load_workbook(FILE_PATH)
    except FileNotFoundError as e:
        print(e)
        return []
    worksheet = workbook.active

    menus_list = []
    for row in range(1, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row, column=col)
            if cell.value is not None:
                if col == 1:
                    menu = {'id': cell.value,
                            'title': worksheet.cell(row=row, column=col + 1).value,
                            'description': worksheet.cell(row=row, column=col + 2).value,
                            'submenus': []}
                    menus_list.append(menu)
                    break
                elif col == 2:
                    submenu = {'id': cell.value,
                               'title': worksheet.cell(row=row, column=col + 1).value,
                               'description': worksheet.cell(row=row, column=col + 2).value,
                               'dishes': []}
                    menus_list[-1]['submenus'].append(submenu)  # adding submenu to menu
                    break
                elif col == 3:
                    dish = {'id': cell.value,
                            'title': worksheet.cell(row=row, column=col + 1).value,
                            'description': worksheet.cell(row=row, column=col + 2).value,
                            'price': str(worksheet.cell(row=row, column=col + 3).value)}
                    menus_list[-1]['submenus'][-1]['dishes'].append(dish)  # adding dish to submenu
                    break
    return menus_list


def update_excel_menu(menus_list: list[MenuDetailsResponse]) -> None:
    '''
    Updates xlsx file with new id's from db
    Walk through file and rewrite corresponding id from menus list
    '''
    try:
        workbook = openpyxl.load_workbook(FILE_PATH)
    except FileNotFoundError as e:
        print(e)
        return
    worksheet = workbook.active

    menu_index = -1
    submenu_index = -1
    dish_index = -1
    for row in range(1, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row, column=col)
            if cell.value is not None:
                if col == 1:
                    menu_index += 1
                    submenu_index = -1
                    dish_index = -1
                    cell.value = menus_list[menu_index]['id']
                    break
                elif col == 2:
                    submenu_index += 1
                    dish_index = -1
                    cell.value = menus_list[menu_index]['submenus'][submenu_index]['id']
                    break
                elif col == 3:
                    dish_index += 1
                    cell.value = menus_list[menu_index]['submenus'][submenu_index]['dishes'][dish_index]['id']
                    break
    workbook.save(FILE_PATH)


async def cache_discounts() -> None:
    '''
    Caches discounts from xlsx file to redis. And invalidate cache for dish and dish list if discount changed
    Don't store None or 0 discount
    '''
    try:
        workbook = openpyxl.load_workbook(FILE_PATH)
    except FileNotFoundError as e:
        print(e)
        return
    worksheet = workbook.active

    menu_id = None
    submenu_id = None
    dish_id = None
    for row in range(1, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row, column=col)
            if cell.value is not None:
                if col == 1:    # menu
                    menu_id = cell.value
                    break
                elif col == 2:  # submenu
                    submenu_id = cell.value
                    break
                elif col == 3:  # dishes
                    dish_id = cell.value
                    old_discount = await cache.get(ce.discount, dish_id=dish_id)
                    new_discount = worksheet.cell(row=row, column=col + 4).value
                    if new_discount != old_discount:
                        if new_discount is None or new_discount == 0:    # delete old_discount, don't save 0 discount
                            await cache.delete(ce.discount, dish_id=dish_id)
                        else:   # set new_discount, overwriting old_discount
                            await cache.set(ce.discount, value=new_discount, dish_id=dish_id)
                        # delete cache for dish_list, dish
                        await cache.delete(ce.dish_list, ce.dish, menu_id=menu_id, submenu_id=submenu_id, dish_id=dish_id)


@async_client
async def get_full_menu(client: AsyncClient) -> list[MenuDetailsResponse]:
    response = await client.get('/api/v1/menus/details')
    return response.json()


@async_client
async def add_menu(menu: dict, client: AsyncClient) -> MenuResponse:
    response = await client.post('/api/v1/menus/', json=menu)
    return response.json()


@async_client
async def update_menu(menu_id: int, data: dict, client: AsyncClient) -> MenuResponse:
    response = await client.patch(f'/api/v1/menus/{menu_id}', json=data)
    return response.json()


@async_client
async def delete_menu(menu_id: int, client: AsyncClient) -> None:
    response = await client.delete(f'/api/v1/menus/{menu_id}')
    return response.json()


@async_client
async def add_submenu(menu_id: int, submenu: dict, client: AsyncClient) -> SubmenuResponse:
    response = await client.post(f'/api/v1/menus/{menu_id}/submenus/', json=submenu)
    return response.json()


@async_client
async def update_submenu(menu_id: int, submenu_id: int, data: dict, client: AsyncClient) -> SubmenuResponse:
    response = await client.patch(f'/api/v1/menus/{menu_id}/submenus/{submenu_id}', json=data)
    return response.json()


@async_client
async def delete_submenu(menu_id: int, submenu_id: int, client: AsyncClient) -> None:
    response = await client.delete(f'/api/v1/menus/{menu_id}/submenus/{submenu_id}')
    return response.json()


@async_client
async def add_dish(menu_id: int, submenu_id: int, dish: dict, client: AsyncClient) -> DishResponse:
    response = await client.post(f'/api/v1/menus/{menu_id}/submenus/{submenu_id}/dishes/', json=dish)
    return response.json()


@async_client
async def update_dish(menu_id: int, submenu_id: int, dish_id: int, data: dict, client: AsyncClient) -> DishResponse:
    response = await client.patch(f'/api/v1/menus/{menu_id}/submenus/{submenu_id}/dishes/{dish_id}', json=data)
    return response.json()


@async_client
async def delete_dish(menu_id: int, submenu_id: int, dish_id: int, client: AsyncClient) -> None:
    response = await client.delete(f'/api/v1/menus/{menu_id}/submenus/{submenu_id}/dishes/{dish_id}')
    return response.json()


async def remove_extra_records(full_menu_db: list[MenuDetailsResponse], full_menu_excel: list[MenuDetailsResponse]) -> None:
    # delete all extra menus/submenus/dishes from db, which not in xls file
    for menu_db in full_menu_db:
        # search menu in xls that equals menu_db by id
        similar_menu_xls_lst: list[MenuDetailsResponse] = list(
            filter(lambda menu_xls: menu_xls['id'] == menu_db['id'], full_menu_excel))
        if not similar_menu_xls_lst:  # delete menu from db if its id not in xls file
            await delete_menu(menu_db['id'])
        else:
            similar_menu_xls: MenuDetailsResponse = similar_menu_xls_lst[0]
            for submenu_db in menu_db['submenus']:
                # search submenu in xls that equals submenu_db by id
                similar_submenu_xls_lst = list(
                    filter(lambda submenu_xls: submenu_xls['id'] == submenu_db['id'], similar_menu_xls['submenus']))
                if not similar_submenu_xls_lst:  # delete submenu from db if its id not in xls file
                    await delete_submenu(menu_db['id'], submenu_db['id'])
                else:
                    similar_submenu_xls: SubmenuDetailsResponse = similar_submenu_xls_lst[0]
                    for dish_db in submenu_db['dishes']:
                        # search dish in xls that equals dish_db by id
                        similar_dish_xls = list(filter(lambda dish_xls: dish_xls['id'] == dish_db['id'],
                                                       similar_submenu_xls['dishes']))
                        if not similar_dish_xls:  # delete dish from db if its id not in xls file
                            await delete_dish(menu_db['id'], submenu_db['id'], dish_db['id'])


async def add_update_records(full_menu_db: list[MenuDetailsResponse], full_menu_excel: list[MenuDetailsResponse]) -> None:
    async def create_menu():
        menu = MenuRequest(title=menu_xls['title'], description=menu_xls['description'])
        menu_xls['id'] = (await add_menu(menu.model_dump()))['id']

    async def create_submenu():
        submenu = SubmenuRequest(title=submenu_xls['title'], description=submenu_xls['description'])
        submenu_xls['id'] = (await add_submenu(menu_xls['id'], submenu.model_dump()))['id']

    async def create_dish():
        dish = DishRequest(title=dish_xls['title'], description=dish_xls['description'], price=dish_xls['price'])
        dish_xls['id'] = (await add_dish(menu_xls['id'], submenu_xls['id'], dish.model_dump()))['id']

    async def update_menu_if_needed():
        # if menu_xls != similar_menu_db (without submenus), then update similar_menu_db
        if {**similar_menu_db, 'submenus': None} != {**menu_xls, 'submenus': None}:
            menu = MenuRequest(title=menu_xls['title'], description=menu_xls['description'])
            await update_menu(menu_xls['id'], menu.model_dump())

    async def update_submenu_if_needed():
        # if submenu_xls != similar_submenu_db (without dishes), then update similar_submenu_db
        if {**similar_submenu_db, 'dishes': None} != {**submenu_xls, 'dishes': None}:
            submenu = SubmenuRequest(title=submenu_xls['title'], description=submenu_xls['description'])
            await update_submenu(menu_xls['id'], submenu_xls['id'], submenu.model_dump())

    async def update_dish_if_needed():
        # if dish_xls != similar_dish_db, then update similar_dish_db
        if similar_dish_db != dish_xls:
            dish = DishRequest(title=dish_xls['title'], description=dish_xls['description'], price=dish_xls['price'])
            await update_dish(menu_xls['id'], submenu_xls['id'], dish_xls['id'], dish.model_dump())

    for menu_xls in full_menu_excel:  # adding/updating all menus, submenus, dishes from excel to db
        if menu_xls not in full_menu_db:  # if this menu (with submenu, dishes) don't exist in db
            # search menu_db which have the same id as menu_xls

            similar_menu_db_lst = [menu_db for menu_db in full_menu_db if menu_db['id'] == menu_xls['id']]
            if similar_menu_db_lst:
                similar_menu_db: MenuDetailsResponse = similar_menu_db_lst[0]
                await update_menu_if_needed()

                for submenu_xls in menu_xls['submenus']:  # adding/updating all submenus from excel to db
                    # search submenu_db which have the same id as submenu_xls
                    similar_submenu_db_lst = [submenu_db for submenu_db in similar_menu_db['submenus']
                                              if submenu_db['id'] == submenu_xls['id']]
                    if similar_submenu_db_lst:
                        similar_submenu_db: SubmenuDetailsResponse = similar_submenu_db_lst[0]
                        await update_submenu_if_needed()

                        for dish_xls in submenu_xls['dishes']:  # adding/updating dish in db
                            # search dish_db which have the same id as dish_xls
                            similar_dish_db = [dish_db for dish_db in similar_submenu_db['dishes'] if
                                               dish_db['id'] == dish_xls['id']]
                            if similar_dish_db:
                                similar_dish_db = similar_dish_db[0]
                                await update_dish_if_needed()
                            else:   # add dishes to db
                                await create_dish()
                    else:   # add submenu, dishes to db
                        await create_submenu()
                        for dish_xls in submenu_xls['dishes']:  # adding dish in db
                            await create_dish()
            else:   # add menu, submenu, dishes to db
                await create_menu()
                for submenu_xls in menu_xls['submenus']:  # adding submenu in db
                    await create_submenu()
                    for dish_xls in submenu_xls['dishes']:  # adding dish in db
                        await create_dish()


async def update_db_from_admin_async():
    full_menu_db = await get_full_menu()
    full_menu_excel = get_excel_menu()

    if full_menu_db != full_menu_excel:     # don't work if different order of menus/submenus/dishes
        await remove_extra_records(full_menu_db, full_menu_excel)
        await add_update_records(full_menu_db, full_menu_excel)
        update_excel_menu(full_menu_excel)
    await cache_discounts()


@celery.task
def update_db_from_admin():
    loop = asyncio.get_event_loop()
    # loop = asyncio.get_event_loop_policy().new_event_loop()
    loop.run_until_complete(update_db_from_admin_async())
    # loop.close()


# update_db_from_admin()
if CELERY_ON:
    celery.add_periodic_task(CELERY_TASK_TIME, update_db_from_admin.s(), name='task-name')
